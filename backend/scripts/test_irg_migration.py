
import sys
import os
import io
import openpyxl
from decimal import Decimal
from sqlalchemy.orm import Session

# Add backend to path
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

from database import SessionLocal, engine, Base
from services.irg_calculator import get_irg_calculator
from models import IRGBareme

def create_mock_excel():
    """Create a mock IRG Excel file in memory"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["MONTANT", "IRG"])
    # Add a few tiers
    data = [
        (30000, 0),
        (35000, 1000),
        (40000, 2500),
        (100000, 15000)
    ]
    for row in data:
        ws.append(row)
    
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

def test_migration_and_calc():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        # 1. Clear existing IRG
        db.query(IRGBareme).delete()
        db.commit()
        
        # 2. Simulate Upload Logic (similar to route)
        print("📥 Simulating Import...")
        mock_file = create_mock_excel()
        wb = openpyxl.load_workbook(mock_file, data_only=True)
        sheet = wb.active
        
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                b = IRGBareme(
                    salaire_min=Decimal(str(row[0])),
                    irg=Decimal(str(row[1])),
                    actif=True
                )
                db.add(b)
                count += 1
        db.commit()
        print(f"✅ Imported {count} rows")
        
        # 3. Test Calculator usage of DB
        print("🧮 Testing Calculator...")
        calc = get_irg_calculator(db)
        # Force reload to be sure
        calc.recharger_bareme()
        
        # Test Case 1: Below threshold
        res1 = calc.calculer_irg(Decimal(29000))
        print(f"Test 29000 -> {res1} (Expected 0)")
        assert res1 == 0
        
        # Test Case 2: Exact match
        res2 = calc.calculer_irg(Decimal(35000))
        print(f"Test 35000 -> {res2} (Expected 1000)")
        assert res2 == 1000
        
        # Test Case 3: Between thresholds (should take previous)
        res3 = calc.calculer_irg(Decimal(38000))
        print(f"Test 38000 -> {res3} (Expected 1000)")
        assert res3 == 1000
        
        print("✅ All Tests Passed!")
        
    except Exception as e:
        print(f"❌ Test Failed: {e}")
        raise
    finally:
        db.close()

if __name__ == "__main__":
    test_migration_and_calc()

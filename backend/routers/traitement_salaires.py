"""
API Router pour le nouveau module Traitement Salaires v3.0
Remplace edition_salaires.py (ancien système)
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from database import get_db
from services.salary_processor import SalaireProcessor
from services.pdf_generator import PDFGenerator
from models import Salaire, Employe, Avance, Credit, StatutCredit, RetenueCredit
from sqlalchemy import inspect

router = APIRouter(prefix="/traitement-salaires", tags=["Traitement Salaires v3.0"])

# Liste des colonnes valides du modèle Salaire (pour filtrage)
SALAIRE_COLUMNS = {c.key for c in inspect(Salaire).mapper.column_attrs}


# ⭐ v3.6.1: Fonctions helper pour suivi des déductions

def _enregistrer_retenues_credits(db: Session, employe_id: int, annee: int, mois: int, resultat: dict):
    """
    Enregistrer les retenues de crédits dans la table retenues_credit
    Met à jour montant_retenu et statut du crédit
    """
    # Récupérer tous les crédits EN_COURS de l'employé
    credits_actifs = db.query(Credit).filter(
        Credit.employe_id == employe_id,
        Credit.statut == StatutCredit.EN_COURS
    ).all()
    
    for credit in credits_actifs:
        # Vérifier si une retenue existe déjà pour ce mois
        retenue_existante = db.query(RetenueCredit).filter(
            RetenueCredit.credit_id == credit.id,
            RetenueCredit.mois == mois,
            RetenueCredit.annee == annee
        ).first()
        
        if not retenue_existante:
            # Créer la retenue
            retenue = RetenueCredit(
                credit_id=credit.id,
                mois=mois,
                annee=annee,
                montant=credit.montant_mensualite,
                date_retenue=date.today()
            )
            db.add(retenue)
            
            # Mettre à jour le cumul
            credit.montant_retenu = Decimal(str(credit.montant_retenu or 0)) + Decimal(str(credit.montant_mensualite))
            
            # Vérifier si le crédit est soldé
            if credit.montant_retenu >= credit.montant_total:
                credit.statut = StatutCredit.SOLDE


def _marquer_avances_deduites(db: Session, employe_id: int, annee: int, mois: int):
    """
    Marquer les avances comme déduites pour le mois donné
    """
    avances = db.query(Avance).filter(
        Avance.employe_id == employe_id,
        Avance.annee_deduction == annee,
        Avance.mois_deduction == mois,
        Avance.deduit == False
    ).all()
    
    for avance in avances:
        avance.deduit = True
        avance.date_deduction = date.today()


@router.get("/preview")
def preview_salaires(
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db)
):
    """
    Calculer salaires de tous les employés actifs (mode brouillon)
    Ne sauvegarde PAS en base de données
    
    Utilisé pour :
    - Prévisualisation avant validation
    - Vérification des calculs
    - Identification des erreurs
    
    Returns:
        Liste de dict avec détails calcul ou erreurs
    """
    processor = SalaireProcessor(db)
    resultats = processor.calculer_tous_salaires(annee, mois)
    
    return {
        "annee": annee,
        "mois": mois,
        "total_employes": len(resultats),
        "success_count": sum(1 for r in resultats if r.get("status") == "OK"),
        "error_count": sum(1 for r in resultats if r.get("status") == "ERROR"),
        "resultats": resultats
    }


@router.get("/preview/{employe_id}")
def preview_salaire_employe(
    employe_id: int,
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    prime_objectif: float = Query(0, ge=0),
    prime_variable: float = Query(0, ge=0),
    db: Session = Depends(get_db)
):
    """
    Calculer salaire d'UN employé (mode brouillon)
    Permet de spécifier primes variables
    
    Returns:
        Dict avec détails complets du calcul
    """
    processor = SalaireProcessor(db)
    resultat = processor.calculer_salaire_employe(
        employe_id=employe_id,
        annee=annee,
        mois=mois,
        prime_objectif=Decimal(str(prime_objectif)),
        prime_variable=Decimal(str(prime_variable))
    )
    
    return resultat


@router.post("/valider/{employe_id}")
def valider_salaire_employe(
    employe_id: int,
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    prime_objectif: float = Query(0, ge=0),
    prime_variable: float = Query(0, ge=0),
    db: Session = Depends(get_db)
):
    """
    Valider et enregistrer le salaire d'un employé en base de données
    
    Actions effectuées :
    1. Calcul complet du salaire
    2. Vérification existence (pas de doublon)
    3. Enregistrement en table salaires
    4. Marquage avances comme déduites
    5. Mise à jour crédits (mois_restants)
    
    Returns:
        Salaire enregistré avec ID
    """
    processor = SalaireProcessor(db)
    
    # 1. Calculer salaire
    resultat = processor.calculer_salaire_employe(
        employe_id=employe_id,
        annee=annee,
        mois=mois,
        prime_objectif=Decimal(str(prime_objectif)),
        prime_variable=Decimal(str(prime_variable))
    )
    
    if resultat.get("status") == "ERROR":
        raise HTTPException(status_code=400, detail=resultat.get("error"))
    
    # 2. Vérifier si salaire existe déjà
    salaire_existant = db.query(Salaire).filter(
        Salaire.employe_id == employe_id,
        Salaire.annee == annee,
        Salaire.mois == mois
    ).first()
    
    if salaire_existant:
        # Mettre à jour
        for key, value in resultat.items():
            if key in SALAIRE_COLUMNS:
                setattr(salaire_existant, key, value)
        
        salaire_existant.statut = "valide"
        salaire_existant.date_paiement = date.today()
        
    else:
        # Créer nouveau salaire - ne garder que les colonnes valides
        salaire_data = {k: v for k, v in resultat.items() if k in SALAIRE_COLUMNS}
        
        salaire_existant = Salaire(**salaire_data)
        salaire_existant.statut = "valide"
        salaire_existant.date_paiement = date.today()
        db.add(salaire_existant)
    
    # 3. Commit
    db.commit()
    db.refresh(salaire_existant)
    
    return {
        "message": "Salaire validé et enregistré avec succès",
        "salaire_id": salaire_existant.id,
        "employe_id": employe_id,
        "salaire_net": str(salaire_existant.salaire_net)
    }


@router.post("/valider-tous")
def valider_tous_salaires(
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db)
):
    """
    Valider et enregistrer TOUS les salaires du mois
    
    Utilise le même processus que valider_salaire_employe
    mais pour tous les employés actifs
    
    ⭐ v3.6.1: Enregistre aussi les retenues (crédits) et marque avances comme déduites
    
    Returns:
        Rapport avec succès/erreurs
    """
    processor = SalaireProcessor(db)
    resultats = processor.calculer_tous_salaires(annee, mois)
    
    success_count = 0
    error_count = 0
    errors = []
    
    for resultat in resultats:
        if resultat.get("status") == "ERROR":
            error_count += 1
            errors.append({
                "employe_id": resultat.get("employe_id"),
                "nom": resultat.get("employe_nom"),
                "error": resultat.get("error")
            })
            continue
        
        try:
            employe_id = resultat["employe_id"]
            
            # Vérifier existence
            salaire_existant = db.query(Salaire).filter(
                Salaire.employe_id == employe_id,
                Salaire.annee == annee,
                Salaire.mois == mois
            ).first()
            
            if salaire_existant:
                # Mettre à jour avec colonnes valides uniquement
                for key, value in resultat.items():
                    if key in SALAIRE_COLUMNS:
                        setattr(salaire_existant, key, value)
                salaire_existant.statut = "valide"
                salaire_existant.date_paiement = date.today()
            else:
                # Créer avec colonnes valides uniquement
                salaire_data = {k: v for k, v in resultat.items() if k in SALAIRE_COLUMNS}
                salaire_existant = Salaire(**salaire_data)
                salaire_existant.statut = "valide"
                salaire_existant.date_paiement = date.today()
                db.add(salaire_existant)
            
            # ⭐ v3.6.1: Enregistrer les retenues de crédits
            _enregistrer_retenues_credits(db, employe_id, annee, mois, resultat)
            
            # ⭐ v3.6.1: Marquer les avances comme déduites
            _marquer_avances_deduites(db, employe_id, annee, mois)
            
            success_count += 1
            
        except Exception as e:
            error_count += 1
            errors.append({
                "employe_id": resultat.get("employe_id"),
                "nom": resultat.get("employe_nom"),
                "error": str(e)
            })
    
    # Commit final
    db.commit()
    
    return {
        "message": f"Validation terminée : {success_count} réussis, {error_count} erreurs",
        "annee": annee,
        "mois": mois,
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors
    }


@router.get("/historique/{employe_id}")
def get_historique_salaires(
    employe_id: int,
    annee: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Récupérer l'historique des salaires d'un employé
    
    Args:
        employe_id: ID employé
        annee: Filtrer par année (optionnel)
    
    Returns:
        Liste des salaires enregistrés
    """
    query = db.query(Salaire).filter(Salaire.employe_id == employe_id)
    
    if annee:
        query = query.filter(Salaire.annee == annee)
    
    salaires = query.order_by(Salaire.annee.desc(), Salaire.mois.desc()).all()
    
    return {
        "employe_id": employe_id,
        "total": len(salaires),
        "salaires": salaires
    }


@router.delete("/supprimer/{salaire_id}")
def supprimer_salaire(
    salaire_id: int,
    db: Session = Depends(get_db)
):
    """
    Supprimer un salaire (si non payé)
    
    Permet d'annuler un salaire validé par erreur
    ATTENTION : Ne restaure PAS les avances/crédits marqués comme déduits
    """
    salaire = db.query(Salaire).filter(Salaire.id == salaire_id).first()
    
    if not salaire:
        raise HTTPException(status_code=404, detail="Salaire non trouvé")
    
    if salaire.statut == "paye":
        raise HTTPException(status_code=400, detail="Impossible de supprimer un salaire déjà payé")
    
    db.delete(salaire)
    db.commit()
    
    return {"message": "Salaire supprimé avec succès", "salaire_id": salaire_id}


@router.get("/statistiques")
def get_statistiques(
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db)
):
    """
    Statistiques globales des salaires du mois
    Calcule depuis preview (brouillon) si aucun salaire validé
    
    Returns:
        Total masse salariale, moyennes, min/max, etc.
    """
    try:
        # Chercher salaires validés
        salaires = db.query(Salaire).filter(
            Salaire.annee == annee,
            Salaire.mois == mois
        ).all()
        
        # Si aucun salaire validé, calculer depuis preview
        if not salaires:
            try:
                processor = SalaireProcessor(db)
                resultats = processor.calculer_tous_salaires(annee, mois)
                
                # Filtrer uniquement les succès
                resultats_ok = [r for r in resultats if r.get("status") == "OK"]
                
                if not resultats_ok:
                    return {
                        "annee": annee,
                        "mois": mois,
                        "nombre_employes": 0,
                        "masse_salariale_brute": "0",
                        "masse_salariale_nette": "0",
                        "masse_cotisable": "0",
                        "masse_imposable": "0",
                        "moyenne_salaire_net": "0",
                        "min_salaire": "0",
                        "max_salaire": "0",
                        "total_irg": "0",
                        "total_securite_sociale": "0"
                    }
                
                salaires_nets = [Decimal(str(r["salaire_net"])) for r in resultats_ok]
                salaires_bruts = [Decimal(str(r["salaire_cotisable"])) for r in resultats_ok]
                salaires_imposables = [Decimal(str(r["salaire_imposable"])) for r in resultats_ok]
                
                return {
                    "annee": annee,
                    "mois": mois,
                    "nombre_employes": len(resultats_ok),
                    "masse_salariale_brute": str(sum(salaires_bruts)),
                    "masse_salariale_nette": str(sum(salaires_nets)),
                    "masse_cotisable": str(sum(salaires_bruts)),
                    "masse_imposable": str(sum(salaires_imposables)),
                    "moyenne_salaire_net": str(sum(salaires_nets) / len(salaires_nets)),
                    "min_salaire": str(min(salaires_nets)),
                    "max_salaire": str(max(salaires_nets)),
                    "total_irg": str(sum(Decimal(str(r["irg"])) for r in resultats_ok)),
                    "total_securite_sociale": str(sum(Decimal(str(r["retenue_securite_sociale"])) for r in resultats_ok))
                }
            except Exception as e:
                # Si erreur, retourner stats vides plutôt que crasher
                print(f"Erreur calcul statistiques preview: {e}")
                return {
                    "annee": annee,
                    "mois": mois,
                    "nombre_employes": 0,
                    "masse_salariale_brute": "0",
                    "masse_salariale_nette": "0",
                    "masse_cotisable": "0",
                    "masse_imposable": "0",
                    "moyenne_salaire_net": "0",
                    "min_salaire": "0",
                    "max_salaire": "0",
                    "total_irg": "0",
                    "total_securite_sociale": "0"
                }
        
        # Si salaires validés existent, utiliser ceux-ci
        salaires_nets = [Decimal(str(s.salaire_net)) for s in salaires]
        salaires_bruts = [Decimal(str(s.salaire_cotisable)) for s in salaires]
        salaires_imposables = [Decimal(str(s.salaire_imposable)) for s in salaires]
        
        return {
            "annee": annee,
            "mois": mois,
            "nombre_employes": len(salaires),
            "masse_salariale_brute": str(sum(salaires_bruts)),
            "masse_salariale_nette": str(sum(salaires_nets)),
            "masse_cotisable": str(sum(salaires_bruts)),
            "masse_imposable": str(sum(salaires_imposables)),
            "moyenne_salaire_net": str(sum(salaires_nets) / len(salaires_nets)),
            "min_salaire": str(min(salaires_nets)),
            "max_salaire": str(max(salaires_nets)),
            "total_irg": str(sum(Decimal(str(s.irg)) for s in salaires)),
            "total_securite_sociale": str(sum(Decimal(str(s.retenue_securite_sociale)) for s in salaires))
        }
    except Exception as e:
        # Fallback complet en cas d'erreur
        print(f"Erreur statistiques: {e}")
        return {
            "annee": annee,
            "mois": mois,
            "nombre_employes": 0,
            "masse_salariale_brute": "0",
            "masse_salariale_nette": "0",
            "masse_cotisable": "0",
            "masse_imposable": "0",
            "moyenne_salaire_net": "0",
            "min_salaire": "0",
            "max_salaire": "0",
            "total_irg": "0",
            "total_securite_sociale": "0"
        }


@router.get("/rapport-pdf")
def generer_rapport_pdf(
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db)
):
    """Générer rapport récapitulatif PDF des salaires du mois"""
    processor = SalaireProcessor(db)
    resultats = processor.calculer_tous_salaires(annee, mois)
    
    # Filtrer uniquement les succès
    resultats_ok = [r for r in resultats if r.get("status") == "OK"]
    
    if not resultats_ok:
        raise HTTPException(status_code=404, detail="Aucun salaire calculé pour cette période")
    
    # Générer PDF
    pdf_gen = PDFGenerator(db)
    pdf_buffer = pdf_gen.generate_rapport_salaires(resultats_ok, {"annee": annee, "mois": mois})
    
    return StreamingResponse(
        BytesIO(pdf_buffer.getvalue()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapport_salaires_{mois}_{annee}.pdf"}
    )


@router.get("/rapport-excel")
def generer_rapport_excel(
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db)
):
    """Générer rapport récapitulatif Excel des salaires du mois"""
    processor = SalaireProcessor(db)
    resultats = processor.calculer_tous_salaires(annee, mois)
    
    # Filtrer uniquement les succès
    resultats_ok = [r for r in resultats if r.get("status") == "OK"]
    
    if not resultats_ok:
        raise HTTPException(status_code=404, detail="Aucun salaire calculé pour cette période")
    
    # Créer workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salaires {mois}-{annee}"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes
    headers = ["Nom", "Prénom", "J. Travaillés", "Salaire Base", "Salaire Cotisable", 
               "Salaire Imposable", "IRG", "Salaire Net"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données
    row = 2
    for r in resultats_ok:
        ws.cell(row=row, column=1, value=r["employe_nom"]).border = border
        ws.cell(row=row, column=2, value=r["employe_prenom"]).border = border
        ws.cell(row=row, column=3, value=r["jours_travailles"]).border = border
        ws.cell(row=row, column=4, value=float(r["salaire_base"])).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(r["salaire_cotisable"])).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=float(r["salaire_imposable"])).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=float(r["irg"])).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=float(r["salaire_net"])).number_format = '#,##0.00'
        row += 1
    
    # Totaux
    row += 1
    ws.cell(row=row, column=1, value="TOTAUX").font = Font(bold=True)
    ws.cell(row=row, column=4, value=sum(float(r["salaire_base"]) for r in resultats_ok)).number_format = '#,##0.00'
    ws.cell(row=row, column=5, value=sum(float(r["salaire_cotisable"]) for r in resultats_ok)).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=sum(float(r["salaire_imposable"]) for r in resultats_ok)).number_format = '#,##0.00'
    ws.cell(row=row, column=7, value=sum(float(r["irg"]) for r in resultats_ok)).number_format = '#,##0.00'
    ws.cell(row=row, column=8, value=sum(float(r["salaire_net"]) for r in resultats_ok)).number_format = '#,##0.00'
    
    # Ajuster largeurs colonnes
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Sauvegarder dans buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rapport_salaires_{mois}_{annee}.xlsx"}
    )


@router.post("/bulletins")
def generer_bulletins_paie(
    data: dict,
    db: Session = Depends(get_db)
):
    """Générer les bulletins de paie pour les employés sélectionnés"""
    annee = data.get("annee")
    mois = data.get("mois")
    employe_ids = data.get("employe_ids", [])
    
    if not employe_ids:
        raise HTTPException(status_code=400, detail="Aucun employé sélectionné")
    
    processor = SalaireProcessor(db)
    pdf_gen = PDFGenerator(db)
    
    bulletins_data = []
    
    for employe_id in employe_ids:
        # Calculer salaire
        resultat = processor.calculer_salaire_employe(employe_id, annee, mois)
        
        if resultat.get("status") != "OK":
            continue
        
        # Récupérer infos employé
        employe = db.query(Employe).filter(Employe.id == employe_id).first()
        if not employe:
            continue
        
        employe_data = {
            "id": employe.id,
            "nom": employe.nom,
            "prenom": employe.prenom,
            "poste_travail": employe.poste_travail,
            "numero_secu_sociale": employe.numero_secu_sociale,
            "date_recrutement": str(employe.date_recrutement),
            "salaire_base": str(employe.salaire_base)
        }
        
        bulletins_data.append({
            "employe_data": employe_data,
            "salaire_data": resultat
        })
    
    if not bulletins_data:
        raise HTTPException(status_code=404, detail="Aucun bulletin généré")
    
    # Générer PDF combiné
    pdf_buffer = pdf_gen.generate_tous_bulletins_combines(
        bulletins_data,
        {"annee": annee, "mois": mois}
    )
    
    return StreamingResponse(
        BytesIO(pdf_buffer.getvalue()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=bulletins_paie_{mois}_{annee}.pdf"}
    )


@router.get("/g29")
def generer_g29(
    annee: int = Query(..., ge=2000, le=2100),
    mois: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db)
):
    """Générer fichier G29 (CNAS) format Excel"""
    processor = SalaireProcessor(db)
    resultats = processor.calculer_tous_salaires(annee, mois)
    
    # Filtrer uniquement les succès
    resultats_ok = [r for r in resultats if r.get("status") == "OK"]
    
    if not resultats_ok:
        raise HTTPException(status_code=404, detail="Aucun salaire calculé pour cette période")
    
    # Créer workbook Excel format G29
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "G29"
    
    # En-têtes G29 (format CNAS)
    headers = ["N°", "N° Sécurité Sociale", "Nom", "Prénom", "Salaire Brut", "Cotisation SS"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
    
    # Données
    for idx, r in enumerate(resultats_ok, 1):
        # Récupérer employé
        employe = db.query(Employe).filter(Employe.id == r["employe_id"]).first()
        
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=employe.numero_secu_sociale if employe else "")
        ws.cell(row=idx+1, column=3, value=r["employe_nom"])
        ws.cell(row=idx+1, column=4, value=r["employe_prenom"])
        ws.cell(row=idx+1, column=5, value=float(r["salaire_cotisable"])).number_format = '#,##0.00'
        ws.cell(row=idx+1, column=6, value=float(r["retenue_securite_sociale"])).number_format = '#,##0.00'
    
    # Totaux
    row = len(resultats_ok) + 2
    ws.cell(row=row, column=4, value="TOTAUX").font = Font(bold=True)
    ws.cell(row=row, column=5, value=sum(float(r["salaire_cotisable"]) for r in resultats_ok)).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=sum(float(r["retenue_securite_sociale"]) for r in resultats_ok)).number_format = '#,##0.00'
    
    # Ajuster largeurs
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=G29_{mois}_{annee}.xlsx"}
    )

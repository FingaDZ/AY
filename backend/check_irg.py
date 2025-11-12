import openpyxl
from pathlib import Path

# Vérifier le fichier à la racine
fichier_racine = Path(__file__).parent.parent / "irg.xlsx"
print(f"Fichier racine: {fichier_racine}")
print(f"Existe: {fichier_racine.exists()}")

if fichier_racine.exists():
    wb = openpyxl.load_workbook(fichier_racine)
    sheet = wb.active
    
    print(f"\nNombre de lignes: {sheet.max_row}")
    print(f"Nombre de colonnes: {sheet.max_column}")
    print(f"En-têtes: {[c.value for c in sheet[1]]}")
    
    print("\n10 premières lignes de données:")
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=11, values_only=True), 1):
        print(f"{i}. Salaire={row[0]}, IRG={row[1]}")
    
    print("\n10 dernières lignes:")
    for i, row in enumerate(sheet.iter_rows(min_row=sheet.max_row-9, max_row=sheet.max_row, values_only=True), 1):
        print(f"{i}. Salaire={row[0]}, IRG={row[1]}")
    
    wb.close()

# Vérifier le fichier dans data/
fichier_data = Path(__file__).parent / "data" / "irg.xlsx"
print(f"\n\nFichier data: {fichier_data}")
print(f"Existe: {fichier_data.exists()}")

if fichier_data.exists():
    wb = openpyxl.load_workbook(fichier_data)
    sheet = wb.active
    print(f"Nombre de lignes: {sheet.max_row}")
    wb.close()

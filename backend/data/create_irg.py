"""
Script pour créer un fichier irg.xlsx avec un barème IRG exemple
Basé sur le barème algérien (à adapter selon la législation en vigueur)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

def creer_fichier_irg():
    """Créer le fichier irg.xlsx avec un barème exemple"""
    
    # Créer un nouveau classeur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barème IRG"
    
    # Style pour l'en-tête
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    ws["A1"] = "Salaire"
    ws["B1"] = "IRG"
    
    ws["A1"].font = header_font
    ws["B1"].font = header_font
    ws["A1"].fill = header_fill
    ws["B1"].fill = header_fill
    ws["A1"].alignment = header_alignment
    ws["B1"].alignment = header_alignment
    
    # Barème IRG exemple (Algérie - à vérifier et adapter)
    # Ces valeurs sont approximatives et doivent être vérifiées
    bareme = [
        (0, 0),
        (10000, 0),
        (15000, 0),
        (20000, 0),
        (25000, 0),
        (30000, 0),
        (35000, 500),
        (40000, 1000),
        (42000, 1300),
        (45000, 1750),
        (48000, 2250),
        (50000, 2500),
        (52000, 2900),
        (55000, 3500),
        (58000, 4200),
        (60000, 4500),
        (65000, 5750),
        (70000, 7000),
        (75000, 8500),
        (80000, 9500),
        (85000, 11000),
        (90000, 12500),
        (95000, 14000),
        (100000, 15500),
        (110000, 19000),
        (120000, 22500),
        (130000, 26500),
        (140000, 30000),
        (150000, 33000),
        (160000, 37000),
        (180000, 44000),
        (200000, 51000),
    ]
    
    # Ajouter les données
    for idx, (salaire, irg) in enumerate(bareme, start=2):
        ws[f"A{idx}"] = salaire
        ws[f"B{idx}"] = irg
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    
    # Formater les nombres
    for row in range(2, len(bareme) + 2):
        ws[f"A{row}"].number_format = '#,##0'
        ws[f"B{row}"].number_format = '#,##0'
    
    # Sauvegarder le fichier
    filepath = Path(__file__).parent / "irg.xlsx"
    wb.save(filepath)
    
    print(f"✅ Fichier IRG créé : {filepath}")
    print(f"   Nombre de tranches : {len(bareme)}")
    print(f"   Salaire minimum : {bareme[0][0]:,} DA")
    print(f"   Salaire maximum : {bareme[-1][0]:,} DA")
    print()
    print("⚠️  IMPORTANT : Vérifiez et ajustez le barème selon la législation en vigueur !")

if __name__ == "__main__":
    creer_fichier_irg()
